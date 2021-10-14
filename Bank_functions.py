import os
from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtCore import QObject, Slot
import csv
import openpyxl
import datetime
from pathlib import Path
import shutil

# from qbosdk import QuickbooksOnlineSDK, QuickbooksOnlineSDKError
# from quickbooks import QuickBooks
# from intuitlib.client import AuthClient


class CsvFunctions(QObject):
    @Slot(str, result=str)
    def get_file(self):
        fnames = QtWidgets.QFileDialog.getOpenFileNames(
            caption="Select all CSV files", filter="CSV Files(*.csv *.CSV)"
        )
        for fname in fnames:
            file = fname
        fnames = fnames[0][0]
        return fnames


class CsvConverter:
    def __init__(self, in_file):
        # File import block
        self.input_file = in_file
        self.path = Path(self.input_file)
        self.file_name = self.path.stem
        self.directory = str(self.path.parent)
        self.new_name = ""

        # create temp directory
        self.temp_directory = ""
        # Workbook and sheet setup
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.num_of_sheets = 0

        #     success bool
        self.is_successful = False

    def run_Main(
        self, save_xl: bool = False, del_orig_file: bool = False
    ):  # PASS IN ARGS for Saving Excel Files, Deleting temp folder
        try:
            # create the temp save dir
            self.make_temp_dir()

            # run the process scripts
            self.csv_conversion()
            self.column_formatting()
            self.worksheet_slice()
            self.column_padding()

            # Save the files
            self.save_csv()
            if save_xl is True:
                self.save_xl_files()
            else:
                pass

            # delete the temp directory
            self.is_successful = True
            if self.is_successful:
                self.clean_up(del_orig_file=del_orig_file)
                return "Program Successfully completed task"
            else:
                return "Error something went wrong along the way"

        except (FileNotFoundError, KeyError, ValueError) as error:
            return f"Error:{ error }"

    def csv_conversion(self):
        with open(self.input_file) as f:
            reader = csv.reader(f, delimiter=",")
            for row in reader:
                self.ws.append(row)

            # find start and end dates
            self.end_date = self.ws[5][0].value
            self.start_date = self.ws[self.ws.max_row][0].value

        self.wb.save(self.temp_directory + "//file.xlsx")

    def column_formatting(self):
        # Delete bank header row
        self.ws.delete_rows(1, 3)

        # Delete value date
        self.ws.delete_cols(2)

        # Delete Balance column
        self.ws.delete_cols(5)

        # store number of rows for ref num
        reference_shift = f"F1:F{self.ws.max_row}"

        # make a new column
        self.ws.insert_cols(2)

        # move ref col to new col
        self.ws.move_range(cell_range=reference_shift, cols=-4)

        # Format Date Cell
        for row in self.ws.iter_rows(max_col=1, min_row=2, max_row=self.ws.max_row):
            for cell in row:
                # aligns date to work in datetime
                date_align = datetime.datetime.strptime(cell.value, "%d %b %Y")

                # formats the date correctly
                formatted_date = date_align.strftime("%m-%d-%Y")

                # replaces the date with the formatted date
                cell.value = formatted_date
        self.wb.save(self.temp_directory + "//file_columnFMT.xlsx")

    def worksheet_slice(self):
        max_lines = 995
        count = 0
        if self.ws.max_row >= max_lines:
            for chunks in range(self.ws.min_row, self.ws.max_row, max_lines):
                count += 1
                self.num_of_sheets = count
                max_chunk = chunks + max_lines
                self.wb.create_sheet(title=f"Worksheet_{count}", index=count)

                # Add headers
                for rows in self.ws.iter_rows(
                    max_row=self.ws.min_row, values_only=True
                ):
                    self.wb.worksheets[count].append(rows)

                # Add content
                for rows in self.ws.iter_rows(
                    min_row=chunks + 1, max_row=max_chunk, values_only=True
                ):
                    self.wb.worksheets[count].append(rows)

            del self.wb["Sheet"]
        else:
            pass
        self.wb.save(self.temp_directory + "//file_slice.xlsx")

    def column_padding(self):
        cell_padding = 8
        for sheet in self.wb.worksheets:
            for cell in range(1, sheet.max_column + 1):
                # print(ws.cell(row=1,column=cell).value)
                if sheet.cell(row=1, column=cell).value.isascii():
                    width = len(self.ws.cell(row=1, column=cell).value) + cell_padding
                    sheet.column_dimensions[chr(64 + cell)].width = width
        self.wb.save(self.temp_directory + "//file_columnP.xlsx")

    def save_csv(self):
        sheet_num = 0

        # counts all the sheets in the workbook
        for sheets in self.wb.worksheets:
            sheet_num += 1

            # create new file name

            start_date_obj = datetime.datetime.strptime(
                self.start_date, "%d %b %Y"
            ).strftime("%m_%d_%Y")
            end_date_obj = datetime.datetime.strptime(
                self.end_date, "%d %b %Y"
            ).strftime("%m_%d_%Y")
            filenameprefix = "AwesomeLight Checking"
            filenamesuffix = ".csv"
            googledrive = Path(
                r"G:\My Drive\Awesome Light business files\Banking transactions\Account CSV\Processed"
            )
            new_name = f"{filenameprefix} {start_date_obj}-{end_date_obj}_{self.wb.sheetnames[sheet_num - 1]}{filenamesuffix}"

            new_file_name = "".join([googledrive.as_posix(), "/", new_name])
            # print(new_file_name)

            # open up the new file
            with open(
                new_file_name, "w", newline=""
            ) as f:  # open('test.csv', 'w', newline="") for python 3
                c = csv.writer(f)

                # write in the data to the new file
                for r in sheets.rows:
                    c.writerow([cell.value for cell in r])

                f.close()

    def make_temp_dir(self):
        self.temp_directory = Path(self.directory + "\\temp\\")
        self.temp_directory.mkdir(parents=True, exist_ok=True)
        self.temp_directory = self.temp_directory.as_posix()
        print(self.temp_directory + "//file.xlsx")

    def del_temp_dir(self):
        print("Deleting Temp Directory!")
        shutil.rmtree(self.temp_directory)
        # os.remove(self.temp_directory)

    def save_xl_files(self):
        self.wb.save(self.directory + f"//{self.file_name}.xlsx")

    def clean_up(self, del_orig_file: bool):
        """
         - If All has gone well this will clean up.
        - If error this will roll it back and clean up
        """
        self.del_temp_dir()
        self.rename_orig_files()
        if del_orig_file:
            self.delete_orig_files()

    def rename_orig_files(self):
        """
        Will rename the original file with assessed at the end.
        If work was done sucessfully
        """
        file = self.path.with_name(f"{self.path.stem}_assessed{self.path.suffix}")

        if not file.exists():
            new_file = self.path.rename(file)
            print(new_file)
            self.path = new_file
            self.file_name = self.path.stem

        print(f"File Renamed! {self.path}")

    def delete_orig_files(self):
        """
        Will delete file if it exist
        """
        print(f"deleting file: {self.path}")
        if self.path.exists():
            os.remove(self.path)


# def quickbooks_online():
#     qbo2 = QuickBooks.auth_client()
#     qbo = QuickbooksOnlineSDK(
#         client_id='',
#         client_secret='',
#         refresh_token='',
#         realm_id='',
#         environment='',
#     )
#     qbo.purchases()


#
# def purchase(paymentType):
#     payload = {
#         "PaymentType": paymentType,
#         "AccountRef": {
#             "name": name,
#             "value": value
#         },
#         "Line": [
#             {
#                 "DetailType": "AccountBasedExpenseLineDetail",
#                 "Amount": amount,
#
#             }
#         ],
#         "CurrencyRefType": cft,
#
#     }

if __name__ == "__main__":
    test_file = Path(
        r"G:\My Drive\Awesome Light business files\Banking transactions\Account CSV\1634170073189.csv"
    )

    def make_test_file():
        if not test_file.is_file():
            with open(test_file.as_posix(), "w") as f:
                f.write("test file")

    make_test_file()
    # file = CsvConverter(test_file)
    # file.rename_orig_files()
    # file.delete_orig_files()
